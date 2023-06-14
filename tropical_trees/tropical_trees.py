#
#   A script that extracts data from a HUGE (60k+ rows, 60+ columns)
#   spreadsheet to help biologists do science.
#   Calculates the average height of the 8 surounding trees around
#   an individual tree, and returns a neat spreadsheet with
#   the individuals label, height, neighbouring average height, 
#   and death status.
#   
#   By Marcus Dahlström 2022
#

from cProfile import label
from itertools import pairwise
from turtle import update
import openpyxl as op
import time as t

###### Script config ######
STARTINGROW = 5        # Which row to start reading from
SHEETNAME = 'Database' # Which sheet to read data from
SAMPSPERTREE = 12      # How many samples are taken for each tree
LABELCOL = 15          # Column 'O', Label
HEIGHTCOL = 27         # Column 'AA', Hfill
DEADCOL = 42           # Column 'AP', Dead. Coincidence that his column is the answer to everything?
LASTROW = 64803        # Last row of the excel sheet

 # Input file path
INPUTFILEPATH = "C:\\Users\\Marcus\\Documents\\Code\\python-playground\\tropical_trees\\Treedata\\treedata.xlsm"
# Output file path
OUTPUTFILEPATH = "C:\\Users\\Marcus\\Documents\\Code\\python-playground\\tropical_trees\\Treedata\\output.xlsx"
  
PROGRESSMOD = 50      # After how many trees should a progress update be printed?

# Column formatting for the output file
OUTSTARTROW = 1         # Starting row
OUTLABELCOL = 1         # Label
OUTSPECIESCOL = 2       # Species
OUTDEADCOL = 3          # Dead status
OUTHEIGHTCOL = 4        # Height
OUTNAVGHEIGHTCOL = 5    # Average neighbouring height
OUTNEIGHTSTARTCOL = 6   # Start of neighbours


###### Tree object ######
class tree:
    def __init__(self, label) -> None:
        # Required: Each tree has a unique label and id.
        # Plot id reoccurs between sites.
        # Three different sites, with 18 different plots, and 100 trees ("positions") in each plot.
        # First character determines site, S, R, M.
        # Second character does not matter.
        # Plot id is K-M and 1-3, one letter and one number.
        # Pos id is A-J and 1-10, one letter and one number.
        # Assume that new trees are alive on init.
        # Assume 8 neighbours, even for corner and edge positions.
        # A unique ID looks like: SxK2D03_xxx
        self.label = label
        self.id = label[0:7]
        self.site = label[0:2]
        self.plot = label[2:4]
        self.pos = label[4:7]
        self.species = label[8:]
        self.height = None
        self.dead = None
        self.neighbours = [None] * 8
        self.avg_height = [] # List of neightbouring trees heights
    
    def update(self, label) -> None:
        self.label = label
        self.id = label[0:7]
        self.site = label[0:2]
        self.plot = label[2:4]
        self.pos = label[4:7]
        self.species = label[8:]
    
    def addAvgHeight(self, height) -> None:
        self.avg_height.append(height)
        
    def getAvgHeight(self) -> float:
        if len(self.avg_height):
            return sum(self.avg_height) / len(self.avg_height)
        else:
            return 0

### Help function for updating the neightbours lists ###
def updateNeighbours(current_tree, neigh_ids) -> None:
    # Update neighbours lists of this tree and neighbouring trees
        for i in range(0,8):
            neigh_id = neigh_ids[i]
            if neigh_id == None:
                continue
            
            neigh_hash = hash(neigh_id)
            # Create new trees from neighbours if they don't exists
            if not neigh_hash in trees.keys():
                trees[neigh_hash] = tree(neigh_id)
                
            # Add the current tree as neighbour to neighbouring trees neighbours lists
            # TODO: this line puts the current tree as the neighbour on the wrong index
            trees[neigh_hash].neighbours[7 - i] = current_tree.id
            
            # Add current tree height to neightbours average height
            if isinstance(current_tree.height, int) or isinstance(current_tree.height, float):
                trees[neigh_hash].addAvgHeight(current_tree.height)
            
            # Add the neighbours to the current tree neighbours list
            if trees[tree_hash].neighbours[i] == None:
                trees[tree_hash].neighbours[i] = neigh_id

###### Main script ######
datawb = op.load_workbook(filename=INPUTFILEPATH, read_only=True, data_only=True)
sheet = datawb[SHEETNAME]
trees = {} # Dict to keep track of all the trees that have been extracted
nr_of_rows = 0 # To keep track of the progress
last_label = None # To keep track of which tree is worked on
tree_is_dead = 0
header_row = 0
last_height = 0.0
first_run = True
    
# Iterate the individual trees and extract the data until no data rows remain
for row in sheet.rows:
    nr_of_rows += 1
    
    # Skip the first 5 rows, as they are just headers
    if header_row < STARTINGROW:
        header_row += 1
        continue
    
    # Start parsing the label into site, plot, and species
    tree_label = row[LABELCOL - 1].value
    # TODO: the whole avg height thing does not work
    if first_run:
        last_label = tree_label
        pass
    elif tree_label != last_label or nr_of_rows == LASTROW: # We're on a new tree.
        current_hash = hash(current_tree.label[0] + current_tree.label[2:7])
        updateNeighbours(trees[current_hash], neighbours_id) # Last thing that is done before staring the computation on the new tree is to update all neighbours of the tree we were on
        last_label = tree_label
        last_height = 0
    
    # Add new trees to the dict
    # or update already existing (created from earlier trees as neighbours)
    # Only the plot and position of the tree are hashed
    # as these are then used for neighbour lookup. Species can't be "calculated" in advanced.
    current_tree = tree(tree_label)
    tree_hash = hash(current_tree.label[0] + current_tree.label[2:7])
    if tree_hash not in trees.keys():
        trees[tree_hash] = current_tree
    else:
        trees[tree_hash].update(tree_label)
        
    # Read the height of the current tree
    if isinstance(row[HEIGHTCOL - 1].value, int) or isinstance(row[HEIGHTCOL - 1].value, float):
        if row[HEIGHTCOL -1].value > last_height:
            last_height = row[HEIGHTCOL - 1].value
            trees[tree_hash].height = last_height
    
    # Check the death status of the current tree
    if row[DEADCOL - 1].value:
        trees[tree_hash].dead = 1
    
    # Add this tree to trees that are its neighbour.
    # Plots:
    #   K1 L1 M1
    #   K2 L2 M2
    #   K3 L3 M3
    #   K4 L4 M4
    #   K5 L5 M5
    #   K6 L6 M6
    # Positions:
    #   A01 B01 C01 ... J01
    #   ...
    #   A10 B10 C10 ... J10
    # Neighbours in plots are looked at clockwise, starting with the top left one
    #   0 1 2
    #   3 x 4
    #   5 6 7
    # Case: Individual in center => 8 neighbours
    # Case: Individual at edge => 5 neighbours
    # Case: Individual in corner => only 3 neighbours
    # Neighbours exits between plots, but not between sites.
    neighbours_plot = [None] * 8
    neighbours_pos= [None] * 8
    plot_chars = ["K", "L", "M"]
    pos_chars = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]

    # Extract plot and pos id from the current tree
    this_plot = [trees[tree_hash].plot[0], int(trees[tree_hash].plot[1])]
    this_pos = [trees[tree_hash].pos[0], int(trees[tree_hash].pos[1:])]

    # Get the char index in the char lists
    this_plot_idx = plot_chars.index(this_plot[0])
    this_pos_idx = pos_chars.index(this_pos[0])

    ### Calculate the neighbouring plots and positions ###
    # Below are four edge cases, where there are no neighbours
    # Corner positions are handled by left and right edges
    # as such the top and bottom edges are more like "middle" positions.
    no_left = no_above = no_right = no_below = False
    if this_plot[0] == "K" and this_pos[0] == "A":
        # This tree is on the left edge => no left neighbours
        no_left = True

    if this_plot[1] == 1 and this_pos[1] == 1:
        # This tree is on the top edge => no above neighbours
        no_above = True

    if this_plot[0] == "M" and this_pos[0] == "J":
        # This tree is on the right edge => no right neighbours
        no_right = True

    if this_plot[1] == 6 and this_pos[1] == 10:
        # This tree is one the bottom edge => no below neighbours
        no_below = True
    
    if not no_left:
        mid_left_neigh_pos = [pos_chars[this_pos_idx - 1], this_pos[1]]
        # Left middle neighbour is in a plot to the left
        if this_pos[0] == "A":
            mid_left_neigh_plot = [plot_chars[this_plot_idx - 1], this_plot[1]]
        else:
            mid_left_neigh_plot = this_plot
        
        neighbours_plot[3] = mid_left_neigh_plot
        neighbours_pos[3] = mid_left_neigh_pos
        
        if not no_above:    
            if this_pos[0] == "A" and this_pos[1] == 1:
                top_left_neigh_plot = [plot_chars[this_plot_idx - 1], this_plot[1] - 1]
                top_left_neigh_pos = ["J", 10]
            elif this_pos[0] == "A":
                top_left_neigh_plot = [plot_chars[this_plot_idx - 1], this_plot[1]]
                top_left_neigh_pos = ["J", this_pos[1] - 1]
            elif this_pos[1] == 1:
                top_left_neigh_plot = [this_plot[0], this_plot[1] - 1]
                top_left_neigh_pos = [pos_chars[this_pos_idx - 1], 10]
            else:
                top_left_neigh_plot = this_plot
                top_left_neigh_pos = [pos_chars[this_pos_idx - 1], this_pos[1] - 1]
            
            neighbours_plot[0] = top_left_neigh_plot
            neighbours_pos[0] = top_left_neigh_pos
            
        if not no_below:
            if this_pos[0] == "A" and this_pos[1] == 10:
                bot_left_neigh_plot = [plot_chars[this_plot_idx - 1], this_plot[1] + 1]
                bot_left_neigh_pos = ["J", 1]
            elif this_pos[0] == "A":
                bot_left_neigh_plot = [plot_chars[this_plot_idx - 1], this_plot[1]]
                bot_left_neigh_pos = ["J", this_pos[1] + 1]
            elif this_pos[1] == 10:
                bot_left_neigh_plot = [this_plot[0], this_plot[1] + 1]
                bot_left_neigh_pos = [pos_chars[this_pos_idx - 1], 1]
            else:
                bot_left_neigh_plot = this_plot
                bot_left_neigh_pos = [pos_chars[this_pos_idx - 1], this_pos[1] + 1]
            
            neighbours_plot[5] = bot_left_neigh_plot
            neighbours_pos[5] = bot_left_neigh_pos
        
    if not no_above:
        # Above neighbour is in a plot above 
        if this_pos[1] == 1:
            mid_above_neigh_pos = [this_pos[0], 10]
            mid_above_neigh_plot = [this_plot[0], this_plot[1] - 1]
        else:
            mid_above_neigh_pos = [this_pos[0], this_pos[1] - 1]
            mid_above_neigh_plot = this_plot
            
        neighbours_plot[1] = mid_above_neigh_plot
        neighbours_pos[1] =  mid_above_neigh_pos
  
    if not no_right:
        # Right middle neighbour is in plot to the right
        if this_pos[0] == "J":
            mid_right_neigh_pos = ["A", this_pos[1]]
            mid_right_neigh_plot = [plot_chars[this_plot_idx + 1], this_plot[1]]
        else:
            mid_right_neigh_pos = [pos_chars[this_pos_idx + 1], this_pos[1]]
            mid_right_neigh_plot = this_plot
        
        neighbours_plot[4] = mid_right_neigh_plot
        neighbours_pos[4] = mid_right_neigh_pos

        if not no_above:
            if this_pos[0] == "J" and this_pos[1] == 1:
                top_right_neigh_plot = [plot_chars[this_plot_idx + 1], this_plot[1] - 1]
                top_right_neigh_pos = ["A", 10]
            elif this_pos[0] == "J":
                top_right_neigh_plot = [plot_chars[this_plot_idx + 1], this_plot[1]]
                top_right_neigh_pos = ["A", this_pos[1] - 1]
            elif this_pos[1] == 1:
                top_right_neigh_plot = [this_plot[0], this_plot[1] - 1]
                top_right_neigh_pos = [pos_chars[this_pos_idx + 1], 10]
            else:
                top_right_neigh_plot = this_plot
                top_right_neigh_pos = [pos_chars[this_pos_idx + 1], this_pos[1] - 1]
                
            neighbours_plot[2] = top_right_neigh_plot
            neighbours_pos[2] = top_right_neigh_pos
            
        if not no_below:
            if this_pos[0] == "J" and this_pos[1] == 10:
                bot_right_neigh_plot = [plot_chars[this_plot_idx + 1], this_plot[1] + 1]
                bot_right_neigh_pos = ["A", 1]
            elif this_pos[0] == "J":
                bot_right_neigh_plot = [plot_chars[this_plot_idx + 1], this_plot[1]]
                bot_right_neigh_pos = ["A", this_pos[1] + 1]
            elif this_pos[1] == 10:
                bot_right_neigh_plot = [this_plot[0], this_plot[1] + 1]
                bot_right_neigh_pos = [pos_chars[this_pos_idx + 1], 1]
            else:
                bot_right_neigh_plot = this_plot
                bot_right_neigh_pos = [pos_chars[this_pos_idx + 1], this_pos[1] + 1]
                
            neighbours_plot[7] = bot_right_neigh_plot
            neighbours_pos[7] = bot_right_neigh_pos
            
    if not no_below:
        # Below neighbour is in a plot below
        if this_pos[1] == 10:
            mid_below_neigh_pos = [this_pos[0], 1]
            mid_below_neigh_plot = [this_plot[0], this_plot[1] + 1]
        else:
            mid_below_neigh_pos = [this_pos[0], this_pos[1] + 1]
            mid_below_neigh_plot = this_plot
        
        neighbours_plot[6] = mid_below_neigh_plot
        neighbours_pos[6] = mid_below_neigh_pos
            
    # Build the neightbours ids.
    # Site is always the same as the current tree.
    neighbours_id = [None] * 8
    for i in range(0, 8):
        if neighbours_plot[i] != None:
            neighbours_id[i] = (current_tree.site[0] + str(neighbours_plot[i][0]) + str(neighbours_plot[i][1]) +
                                str(neighbours_pos[i][0]) + str(neighbours_pos[i][1]).rjust(2, "0"))
    
    #print(f"{trees[tree_hash].site} {trees[tree_hash].plot} {trees[tree_hash].pos} {trees[tree_hash].species} {trees[tree_hash].dead} {trees[tree_hash].height}")
    # Progress update
    if nr_of_rows%PROGRESSMOD == 0:
        print(f"{round(nr_of_rows / SAMPSPERTREE)} trees extracted...")
        
    first_run = False
    
    
    
# Close the data workbook
datawb.close()

## Save the calculated data into a new spreadsheet
outputwb = op.Workbook()
wb = outputwb.active
wb.title = "Tree neighbours"

# Format the column labels
wb.cell(row=OUTSTARTROW, column=OUTLABELCOL, value="Label")
wb.cell(row=OUTSTARTROW, column=OUTSPECIESCOL, value="Species")
wb.cell(row=OUTSTARTROW, column=OUTDEADCOL, value="Dead")
wb.cell(row=OUTSTARTROW, column=OUTHEIGHTCOL, value="Height")
wb.cell(row=OUTSTARTROW, column=OUTNAVGHEIGHTCOL, value="NeighAvgHeight")
neighbour_order = ["TopLeft", "TopCent", "TopRight",
                   "MidLeft", "MidRight",
                   "BotLeft", "BotCent", "BotRight"]
for i in range(0,8):
    wb.cell(row=OUTSTARTROW, column=OUTNEIGHTSTARTCOL + i, value=neighbour_order[i])
    

r = 1
for k, v in trees.items():
    wb.cell(row=OUTSTARTROW + r, column=OUTLABELCOL, value=v.label)
    wb.cell(row=OUTSTARTROW + r, column=OUTSPECIESCOL, value=v.species)
    wb.cell(row=OUTSTARTROW + r, column=OUTDEADCOL, value=v.dead)
    wb.cell(row=OUTSTARTROW + r, column=OUTHEIGHTCOL, value=v.height)
    wb.cell(row=OUTSTARTROW + r, column=OUTNAVGHEIGHTCOL, value=v.getAvgHeight())
    
    # Add neighbours to their right columns
    nc = 0
    for n in v.neighbours:
        wb.cell(row=OUTSTARTROW + r, column=OUTNEIGHTSTARTCOL + nc, value=n)
        nc += 1
    
    r += 1
    # Progress update
    if r%PROGRESSMOD == 0:
        print(f"{r} trees saved...")


outputwb.save(filename=OUTPUTFILEPATH)

print(f"Done. Total of {round(nr_of_rows / SAMPSPERTREE)} trees were extracted. Phew..!")